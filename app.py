import streamlit as st
import pandas as pd
import fitz
from PIL import Image
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="Saisie assistée pressiométrique", layout="wide")
st.title("Saisie assistée des coupes pressiométriques vers Excel")

uploaded = st.file_uploader("Importer le PDF", type=["pdf"])

PROFONDEURS = [1.5, 3, 4.5, 6, 7.5, 9, 10.5, 12, 13.5, 15]


def fr(x):
    return str(x).replace(".", ",")


def pdf_page_to_image(doc, page_index):
    page = doc[page_index]
    pix = page.get_pixmap(matrix=fitz.Matrix(2.5, 2.5))
    return Image.open(BytesIO(pix.tobytes("png")))


def create_empty_rows(page_number):
    rows = []

    for i, p in enumerate(PROFONDEURS):
        rows.append({
            "Page PDF": page_number,
            "Nom du sondages": "" if i != 0 else "",
            "x": "" if i != 0 else "",
            "y": "" if i != 0 else "",
            "Profondeur (m)": fr(p),
            "Lithologie": "",
            "Pl* (MPa)": "",
            "Em (MPa)": ""
        })

    return rows


def make_excel(df):
    output = BytesIO()

    export = df[
        [
            "Nom du sondages",
            "x",
            "y",
            "Profondeur (m)",
            "Lithologie",
            "Pl* (MPa)",
            "Em (MPa)"
        ]
    ]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, startrow=1, sheet_name="Pressiometrique")
        ws = writer.book["Pressiometrique"]

        ws.merge_cells("A1:D1")
        ws["A1"] = "Echantillon"

        ws.merge_cells("E1:G1")
        ws["E1"] = "Caracteristiques pressiometriques"

        blue = PatternFill("solid", fgColor="7EC8E3")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

        for cell in ws[1]:
            cell.fill = blue
            cell.font = Font(bold=True)

        for cell in ws[2]:
            cell.fill = blue
            cell.font = Font(bold=True)

        widths = {
            "A": 18,
            "B": 15,
            "C": 15,
            "D": 16,
            "E": 35,
            "F": 14,
            "G": 14
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    output.seek(0)
    return output


if uploaded:
    pdf_bytes = uploaded.read()
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    st.success(f"PDF chargé : {len(doc)} pages")

    page_index = st.number_input(
        "Choisir la page à afficher",
        min_value=1,
        max_value=len(doc),
        value=1
    ) - 1

    if "data" not in st.session_state:
        all_rows = []
        for p in range(len(doc)):
            all_rows.extend(create_empty_rows(p + 1))
        st.session_state.data = pd.DataFrame(all_rows)

    col1, col2 = st.columns([1.1, 1.4])

    with col1:
        st.subheader(f"Page {page_index + 1}")
        image = pdf_page_to_image(doc, page_index)
        st.image(image, use_container_width=True)

    with col2:
        st.subheader("Tableau à remplir / corriger")

        df_page = st.session_state.data[
            st.session_state.data["Page PDF"] == page_index + 1
        ].copy()

        edited_page = st.data_editor(
            df_page,
            use_container_width=True,
            num_rows="fixed",
            hide_index=True
        )

        if st.button("Enregistrer cette page"):
            st.session_state.data.loc[
                st.session_state.data["Page PDF"] == page_index + 1,
                :
            ] = edited_page.values

            st.success("Page enregistrée.")

    st.divider()

    st.subheader("Aperçu global")
    st.dataframe(st.session_state.data, use_container_width=True)

    excel = make_excel(st.session_state.data)

    st.download_button(
        "Télécharger Excel final",
        data=excel,
        file_name="extraction_pressiometrique.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
